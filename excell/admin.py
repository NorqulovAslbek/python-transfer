from .check_field import check_status, check_phone, check_expire, check_card, clean_balance_for_bigint
from django.contrib import admin
from django.urls import path
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.admin import SimpleListFilter
from openpyxl import load_workbook
import pandas as pd
from .models import Card
from .forms import ExcelImportForm
from django.db import models

from .send_sms import send_sms


class StatusFilter(SimpleListFilter):
    title = 'Status'
    parameter_name = 'status'

    def lookups(self, request, model_admin):
        return (
            ('active', 'Active'),
            ('inactive', 'Inactive'),
            ('expired', 'Expired'),
        )

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(status=self.value())
        return queryset


@admin.register(Card)
class CardAdmin(admin.ModelAdmin):
    list_display = ('card_number', 'expire', 'phone', 'status', 'balance')
    change_list_template = "admin/excel/change_list.html"  # Maxsus shablon
    list_filter = (StatusFilter,)
    search_fields = ['status']
    actions = ['show_selected_cards', ]

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel), name='card-import-excel'),
            path('export-excel/', self.admin_site.admin_view(self.export_excel), name='card-export-excel'),  # Yangi URL
        ]
        return custom_urls + urls

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['show_import_button'] = True
        return super().changelist_view(request, extra_context=extra_context)

    def import_excel(self, request):
        if request.method == 'POST':
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['file']
                wb = load_workbook(excel_file)
                ws = wb.active

                for row in ws.iter_rows(min_row=1):
                    print("\n\n>>>>>>>>>>>>>", row, "\n\n")
                    if row[0].value == "id":
                        continue

                    card_number = str(row[0].value)
                    expire = str(row[1].value)
                    phone = str(row[2].value)
                    status = str(row[3].value)
                    balance_str = str(row[4].value)
                    if card_number and Card.objects.filter(card_number=card_number).exists():
                        self.message_user(
                            request,
                            f"Karta raqami {card_number} allaqachon bazada mavjud! Ushbu qator o'tkazib yuborildi.",
                            level='warning'
                        )
                        continue
                    Card.objects.create(
                        card_number=check_card(card_number),
                        expire=check_expire(expire),
                        phone=check_phone(phone),
                        status=check_status(status),
                        balance=clean_balance_for_bigint(balance_str)
                    )
                self.message_user(request, "Excel file muvaffaqiyatli yuklandi!")
                return redirect("..")
        else:
            form = ExcelImportForm()

        context = {
            "form": form,
            "title": "Import Excel File",
            "app_label": "excel",
            "opts": self.model._meta,
        }
        return render(request, "admin/excel_form.html", context)

    def export_excel(self, request):
        queryset = Card.objects.all()

        status = request.GET.get('status')

        if status:
            queryset = queryset.filter(status=status)
        else:
            # Agar status filtri tanlanmagan bo'lsa, barcha ma'lumotlarni eksport qilamiz
            self.message_user(request, "Iltimos, status bo'yicha filtr tanlang!", level='warning')

        print("\n\n>>>>>>>>>>>>>>", queryset, status, request.GET, "<<<<<<<<<<<<<<<<<\n\n")

        # Qidiruv parametri (agar qo'llanilgan bo'lsa)
        search_query = request.GET.get('q')
        if search_query:
            queryset = queryset.filter(
                models.Q(card_number__icontains=search_query) |
                models.Q(phone__icontains=search_query) |
                models.Q(status__icontains=search_query)
            )

        # Agar filtrlangan ma'lumotlar bo'sh bo'lsa, xabar ko'rsatamiz
        if not queryset.exists():
            self.message_user(request, "Tanlangan filtrlar bo'yicha ma'lumotlar topilmadi!", level='warning')
            return redirect("..")

        # Eksport qilish
        cards = queryset.values('id', 'card_number', 'expire', 'phone', 'status', 'balance')
        df = pd.DataFrame(list(cards))
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = f'attachment; filename="cards_export_{status or "all"}_{pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        df.to_excel(response, index=False, engine='openpyxl')
        return response

    # Yangi action: Tanlangan qatorlarni ro'yxat sifatida ko'rsatish
    def show_selected_cards(self, request, queryset):
        if not queryset.exists():
            self.message_user(request, "Hech qanday qator tanlanmadi!", level='warning')
            return

        # Tanlangan qatorlarni ro'yxat sifatida shakllantirish
        selected_cards = [
            f"Card ID: {card.id}, Number: {card.card_number}, Status: {card.status}, Balance: {card.balance}"
            for card in queryset
        ]
        cards_list = "\n".join(selected_cards)

        # Foydalanuvchiga xabar ko'rsatish
        self.message_user(request, f"Tanlangan qatorlar:\n{cards_list}", level='info')

    show_selected_cards.short_description = "Tanlangan kartalarni ro'yxat sifatida ko'rsatish"

    def show_selected_cards(self, request, queryset):
        if not queryset.exists():
            self.message_user(request, "Hech qanday qator tanlanmadi!", level='warning')
            return
        # Bu yerda sms yuborganma
        for card in queryset:
            phone = card.phone
            balance = card.balance
            card_number = card.card_number
            if phone != "None" and card_number:
                send_sms(balance, phone)

        context = {
            'cards': queryset,
            'opts': self.model._meta,
        }
        return render(request, 'admin/excel/selected_cards.html', context)

    def get_actions(self, request):
        actions = super().get_actions(request)
        if not request.user.is_superuser:
            # Faqat superuser uchun ko'rinadi
            if 'show_selected_cards' in actions:
                del actions['show_selected_cards']
        return actions

    def export_selected_cards(self, request, queryset):
        if not queryset.exists():
            self.message_user(request, "Hech qanday qator tanlanmadi!", level='warning')
            return
        cards = queryset.values('id', 'card_number', 'expire', 'phone', 'status', 'balance')
        df = pd.DataFrame(list(cards))
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = f'attachment; filename="selected_cards_{pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        df.to_excel(response, index=False, engine='openpyxl')
        return response

    export_selected_cards.short_description = "Tanlangan kartalarni Excel'ga eksport qilish"

    actions = ['show_selected_cards', 'export_selected_cards']
